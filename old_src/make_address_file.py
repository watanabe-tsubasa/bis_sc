import openpyxl
import pandas as pd

data_path = './data/company_address.xlsx'
result_path = './out/address_list.xlsx'
company_list = ['北関東', '南関東', '北陸信越', '東海', '近畿', '中四国']

book = openpyxl.Workbook()
book.save(result_path)

def make_address_dataframe(company_name, file_path):
    work_book = openpyxl.load_workbook(file_path)
    temp_sheet = work_book[company_name]
    temp_data = temp_sheet['A1'].value
    # print(temp_data)
    temp_values = temp_data.split(';')
    # print(temp_values)
    
    df = pd.DataFrame(columns=['name', '名前', 'address'])
    for val in temp_values:
        # print(val)
        val = val.split('(')
        val = [val[0], *val[1].split(')')]
        for i, j in enumerate(val):
            val[i] = j.strip()
        # print(val)
        # print(df.columns)
        add_df = pd.DataFrame(val, index=df.columns)
        # print(f'add_df:{add_df}')
        df = pd.concat([df,add_df.T], axis=0)
    df = df.set_index('name')
    
    return df
    
for company in company_list:
    print(company)
    df = make_address_dataframe(company, data_path)
    with pd.ExcelWriter(result_path, mode='a', engine='openpyxl') as writer:
        
        df.to_excel(writer, sheet_name=company)
        
work_book = openpyxl.load_workbook(result_path)
sheets = work_book.sheetnames
# print(sheets[0])
work_book.remove(work_book[sheets[0]])
work_book.save(result_path)