from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import sys
import os
import pandas as pd
import openpyxl

# path一覧
driver_path = r'C:\Users\341137\src\driver\chromedriver.exe'
access_URL = r'https://aeonpeopleportal-web.aeonpeople.biz/employee-search/'
auth_id = '341137'
auth_pass = '3682tw'
data_path = '../out/address_list.xlsx'
result_path = '../out/employee_list.xlsx'

# カンパニーのリスト作成
wb = openpyxl.load_workbook(data_path)
company_list = wb.sheetnames

# アクセスするページのブラウザを立ち上げ
driver = webdriver.Chrome(driver_path)
driver.implicitly_wait(5)
wait = WebDriverWait(driver, 5)
driver.get(access_URL)
title = driver.title
print(title)

# 各店長の名前をシートから抜き出し、苗字と名前とidのリストに整形する
def make_list(sheet_path, company):
    def make_id(address):
        id = address.split('@')[0]
        id = id.replace('<', '')
        return id
    df = pd.read_excel(sheet_path, sheet_name=company)
    address_list = df['address']
    id_list = list(map(make_id, address_list))
    name_list = df['名前'].tolist()
    name_list = list(map(lambda x: x.split(' '), name_list))
    employee_list = []
    for i,j in zip(id_list, name_list):
        employee_list.append([i,*j])
    return employee_list

# 入力されたカンパニーのdataframeを作成する
def make_employees_dataframe(employee_list):
    base_list = []
    
    for id, last_name, first_name in employee_list:
        # driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[1]/div[1]/div/div/div/input').send_keys('イオンリテール（株）')
        driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[2]/div[1]/div/div/div/input').send_keys(id)
        driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[2]/div[2]/div[1]/div/div/div/input').send_keys(last_name)
        driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[2]/div[2]/div[2]/div/div/div/input').send_keys(first_name)
        driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[2]/button[2]').click()
        # time.sleep(0.5)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'u-omitpipe__contents')))

        # HTMLの取得
        print('getting_html')
        html = driver.page_source.encode('utf-8')
        soup = BeautifulSoup(html, 'html.parser')
        spans = soup.find_all('span', class_='u-omitpipe__contents')
        employee_data = list(map(lambda x: x.get_text(), spans))
        
        if len(employee_data) == 6:
            employee_data = employee_data[-2:]
        else:
            try:
                last_idx = employee_data.index('店長')
                employee_data = employee_data[last_idx - 1:last_idx + 1]
            except:
                employee_data = employee_data[-2:]

        print(employee_data)
        base_list.append(employee_data)
        
        # print('success')

        # driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[1]/div[1]/div/div/div/input').click()
        # driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[1]/div[1]/div/div/div/input').sendKeys(Keys.CONTROL + "a")
        # driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[1]/div[1]/div/div/div/input').sendKeys(Keys.DELETE)
        driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[2]/div[1]/div/div/div/input').clear()
        driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[2]/div[2]/div[1]/div/div/div/input').clear()
        driver.find_element(By.XPATH, '//*[@id="employeesearchform"]/div/div/div[1]/div[2]/div[2]/div[2]/div/div/div/input').clear()
     
    df = pd.DataFrame(base_list,columns=['店名', '役職'])    
    return df

# 従業員検索へのログイン
driver.find_element(By.CSS_SELECTOR,'#signin > div > div > div.o-whole__body > section > div > section > div > div > button').click()

driver.find_element(By.CSS_SELECTOR, '#userNameInput').send_keys(auth_id)
driver.find_element(By.CSS_SELECTOR, '#passwordInput').send_keys(auth_pass)
driver.find_element(By.CSS_SELECTOR, '#submitButton').click()

driver.find_element(By.CSS_SELECTOR,'#work-menu > div > div > div.o-whole__body > section:nth-child(3) > div > div.o-grid.o-grid--gutter-pc-h-16.o-grid--gutter-sp-h-16.o-grid--gutter-pc-v-24.o-grid--gutter-sp-v-16.o-grid--favoritetool > div > div:nth-child(17) > a').click()

# 従業員検索

book = openpyxl.Workbook()
book.save(result_path)

with pd.ExcelWriter(result_path, mode='a') as writer:
    for company in company_list:
        employee_list = make_list(sheet_path=data_path, company=company)
        # print(employee_list)
        df = make_employees_dataframe(employee_list=employee_list)
        df_base = pd.read_excel('../out/address_list.xlsx', sheet_name=company)
        df = pd.concat([df_base, df], axis=1)
        df.to_excel(writer, sheet_name=company)

work_book = openpyxl.load_workbook(result_path)
sheets = work_book.sheetnames
# print(sheets[0])
work_book.remove(work_book[sheets[0]])
work_book.save(result_path)

driver.quit()